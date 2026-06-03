import json
import re
import shutil
import sys
from collections import Counter
from pathlib import Path

import openpyxl


SOURCE_COLUMNS = {
    "id": 0,
    "name": 1,
    "technique": 2,
    "flavor": 3,
    "servings": 4,
    "ingredientsText": 5,
    "method": 6,
    "fdcText": 7,
}


def text(value):
    return str(value or "").strip()


def split_semicolon(value):
    return [part.strip() for part in re.split(r"[；;]", text(value)) if part and part.strip()]


def parse_ingredients(value):
    ingredients = []
    for raw in split_semicolon(value):
        match = re.match(r"(.+?)(\d+(?:\.\d+)?\s*(?:kg|g|mg|克|千克|斤|ml|毫升|升|个|颗|根|片|勺|匙|茶匙|汤匙|杯|碗|份))?$", raw, re.I)
        if not match:
            ingredients.append({"name": raw, "amount": ""})
            continue
        name = text(match.group(1))
        amount = text(match.group(2))
        ingredients.append({"name": name, "amount": amount})
    return ingredients


def parse_fdc_matches(value):
    matches = []
    for raw in split_semicolon(value):
        if ":" not in raw:
            continue
        name, fdc_id = raw.split(":", 1)
        matches.append({"name": text(name), "fdcId": text(fdc_id)})
    return matches


def compact_method(value):
    return re.sub(r"\s+", " ", text(value))


def parse_calorie_estimates(source_path):
    estimates = {}
    pattern = re.compile(
        r"^\s*(?P<source_id>\d+)\.\s*"
        r"(?P<name>.+?)（[^）]+）：整道约(?P<total>\d+)大卡，"
        r"每人约(?P<per_serving>\d+)大卡。(?P<detail>.+)$"
    )

    for line in Path(source_path).read_text(encoding="utf-8-sig").splitlines():
        match = pattern.match(line.strip())
        if not match:
            continue
        name = text(match.group("name"))
        estimates[name] = {
            "sourceId": int(match.group("source_id")),
            "totalKcal": int(match.group("total")),
            "perServingKcal": int(match.group("per_serving")),
            "text": compact_method(match.group("detail")),
        }

    return estimates


def attach_calorie_estimates(index, estimate_path):
    estimates = parse_calorie_estimates(estimate_path)
    matched_count = 0
    unmatched_recipe_names = []

    for item in index["items"]:
        estimate = estimates.get(item["name"])
        if not estimate:
            unmatched_recipe_names.append(item["name"])
            continue
        item["calorieEstimate"] = estimate
        item["searchText"] = " ".join([
            item["searchText"],
            str(estimate["totalKcal"]),
            str(estimate["perServingKcal"]),
            "kcal",
            "大卡",
        ]).lower()
        matched_count += 1

    index["calorieEstimateSource"] = Path(estimate_path).name
    index["calorieEstimateCount"] = len(estimates)
    index["calorieEstimateMatchedCount"] = matched_count
    index["calorieEstimateMissingRecipeCount"] = len(unmatched_recipe_names)
    if unmatched_recipe_names:
        index["calorieEstimateMissingRecipes"] = unmatched_recipe_names[:20]


def build_index(source_path, calorie_estimate_path=None):
    workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        next(rows, None)

        items = []
        ingredient_counter = Counter()
        technique_counter = Counter()
        flavor_counter = Counter()

        for row in rows:
            if not row or not text(row[SOURCE_COLUMNS["name"]]):
                continue

            ingredients = parse_ingredients(row[SOURCE_COLUMNS["ingredientsText"]])
            fdc_matches = parse_fdc_matches(row[SOURCE_COLUMNS["fdcText"]])
            technique = text(row[SOURCE_COLUMNS["technique"]])
            flavor = text(row[SOURCE_COLUMNS["flavor"]])
            item = {
                "id": int(row[SOURCE_COLUMNS["id"]] or len(items) + 1),
                "name": text(row[SOURCE_COLUMNS["name"]]),
                "technique": technique,
                "flavor": flavor,
                "servings": text(row[SOURCE_COLUMNS["servings"]]),
                "ingredientsText": text(row[SOURCE_COLUMNS["ingredientsText"]]),
                "ingredients": ingredients,
                "method": compact_method(row[SOURCE_COLUMNS["method"]]),
                "fdcMatches": fdc_matches,
            }
            item["searchText"] = " ".join([
                item["name"],
                item["technique"],
                item["flavor"],
                item["ingredientsText"],
                " ".join(match["fdcId"] for match in fdc_matches),
            ]).lower()
            items.append(item)

            technique_counter[technique] += 1
            flavor_counter[flavor] += 1
            for ingredient in ingredients:
                ingredient_counter[ingredient["name"]] += 1

        index = {
            "source": Path(source_path).name,
            "sheet": worksheet.title,
            "itemCount": len(items),
            "facets": {
                "techniques": [name for name, _count in technique_counter.most_common() if name],
                "flavors": [name for name, _count in flavor_counter.most_common() if name],
                "topIngredients": [name for name, _count in ingredient_counter.most_common(40) if name],
            },
            "items": items,
        }
        if calorie_estimate_path:
            attach_calorie_estimates(index, calorie_estimate_path)
        return index
    finally:
        workbook.close()


def main():
    if len(sys.argv) not in (3, 4):
        print("Usage: build_menu_library_rag.py <source.xlsx> <output.json> [calorie-estimates.txt]", file=sys.stderr)
        return 2

    source_arg = Path(sys.argv[1])
    source_path = source_arg
    if not source_path.exists():
        print(f"Source file not found: {source_path}", file=sys.stderr)
        return 1

    output_path = Path(sys.argv[2])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    calorie_arg = Path(sys.argv[3]) if len(sys.argv) == 4 else None
    if calorie_arg and not calorie_arg.exists():
        print(f"Calorie estimate file not found: {calorie_arg}", file=sys.stderr)
        return 1

    # openpyxl can struggle with non-ASCII paths on some Windows shells, so parse a local ASCII temp copy.
    temp_path = output_path.parent / "_menu_library_source.xlsx"
    shutil.copyfile(source_path, temp_path)
    try:
        index = build_index(temp_path, calorie_arg)
        index["source"] = source_arg.name
    finally:
        temp_path.unlink(missing_ok=True)

    if calorie_arg and index.get("calorieEstimateMatchedCount") != index["itemCount"]:
        missing = index.get("calorieEstimateMissingRecipes", [])
        print(
            f"Calorie estimates matched {index.get('calorieEstimateMatchedCount')} "
            f"of {index['itemCount']} recipes by name. Missing examples: {missing}",
            file=sys.stderr,
        )
        return 1

    output_path.write_text(json.dumps(index, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Wrote {index['itemCount']} recipes to {output_path}")


if __name__ == "__main__":
    raise SystemExit(main())
